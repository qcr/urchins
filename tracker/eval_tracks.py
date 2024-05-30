#!/usr/bin/env/python3

"""Script to numerically evaluate the performance of a YOLO tracker"""

import os
import glob
import cv2 as cv
import numpy as np
import math
import yaml
from ultralytics import YOLO
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import pandas as pd
import openpyxl
import sys


classes = ['Tripys', 'Rings']
green = [0, 255, 00] #Rings
red = [0, 0, 255] #Tripys
class_colours = [green, red]

def read_config(config_file):
    """Open and read the configuration file.
    Args:
        config_file (str): Path to the configuration file.
    Returns:
        dict: Configuration parameters read from the file.
    """
    if not os.path.isfile(config_file):
        print(f"Error: Configuration file '{config_file}' not found.")
        return None
    try:
        with open(config_file, 'r') as file:
            config = yaml.safe_load(file)
    except yaml.YAMLError as e:
        print(f"Error: Failed to parse YAML file '{config_file}': {e}")
        return None
    return config

def save_track_img(results, track_history, track_class, frame, save_path, frame_no, SHOW=False):
    """Save the frame with tracking lines and bounding boxes.
    Args:
        results (list): List of detection results from model.track().
        track_history (dict): Dictionary of track histories.
        track_class (dict): Dictionary of track classes.
        frame (np.ndarray): Frame image.
        save_path (str): Path to save the image.
        frame_no (int): Frame number
        SHOW (bool, optional): Whether to display the image. Defaults to False.
    Returns: True"""
    #try:
    if results[0].boxes.id == None:
        print(f'\rNo track ids found in frame: {frame_no}', end='')
        sys.stdout.flush()
        return track_class
    track_ids = results[0].boxes.id.int().cpu().tolist()
    boxes = results[0].boxes
    for i, track_id in enumerate(track_ids):
        box = boxes[i]
        x, y, w, h = box.xywh[0]
        x1 = int(x - w/2)
        y1 = int(y - h/2)
        class_id = int(box.cls.item())
        conf = box.conf.item()
        track = track_history[track_id]
        track.append((float(x), float(y)))
        tr_cls = track_class[track_id]
        tr_cls.append(class_id)
        if len(track) > 90:  # Retain 90 tracks for 90 frames
            track.pop(0)
        # Draw the tracking lines
        points = np.hstack(track).astype(np.int32).reshape((-1, 1, 2))
        cv.polylines(frame, [points], isClosed=False, color=class_colours[class_id], thickness=10)  # Draw tracking lines
        cv.rectangle(frame, (x1, y1), (int(x+w/2), int(y+h/2)), class_colours[class_id], 2)  # Rectangle around object
        cv.putText(frame, f"{track_id} {classes[class_id]}: {conf:.2f}", (int(x1-5), int(y1-5)), 
                    cv.FONT_HERSHEY_SIMPLEX, fontScale=3.5, color=class_colours[class_id], thickness=3)
    if SHOW:
        cv.imshow("Tracking", frame)
        cv.waitKey(1)
    cv.imwrite(save_path, frame)
    # import code
    # code.interact(local=dict(globals(), **locals()))
    return track_class
    
def class_counts(track_class, class_count_list):
    """From tracks, calcuates the most common class and then returns the count of each class
    Args: 
        track_class (dict): Dictionary of track classes.
        class_count_list (list): List to store the count of each class
    Returns:
        list: List of counts of each class
    """
    for key in track_class.keys():
        length = len(track_class[key])
        total = sum(track_class[key])
        if total/length > 0.5:
            class_count_list[1] += 1
        else:
            class_count_list[0] += 1
    return class_count_list


def track(video_path, save_dir, model, track_config, conf_threshold, frame_skip, max_count=9999999, SAVE=True):
    """Track objects in a video using the YOLOv8 model.
    Args:
        video_path (str): Path to the video file.
        save_dir (str): Directory to save the output images.
        model (YOLO): YOLOv8 model object.
        track_config (str): Path to the tracking configuration file.
        conf_threshold (float): Detection confidence threshold.
        frame_skip (int): Number of frames to skip between detections.
        max_count (int, optional): Maximum number of frames to process. Defaults to 9999999.
        SAVE (bool, optional): Save the individual images
    Returns:
        int: Number of tracks detected.
    """
    # Open the video file
    cap = cv.VideoCapture(video_path)
    video_filename = os.path.splitext(os.path.basename(video_path))[0]
    output_dir = os.path.join(save_dir, video_filename+'_detect')
    os.makedirs(output_dir, exist_ok=True)
    # Define the codec and create VideoWriter object
    codec = cv.VideoWriter_fourcc(*'mp4v')
    output_file = os.path.join(output_dir, f"{video_filename}_BS_trac_demo.mp4")
    output_width = int(cap.get(cv.CAP_PROP_FRAME_WIDTH))
    output_height = int(cap.get(cv.CAP_PROP_FRAME_HEIGHT))
    fps = int(cap.get(cv.CAP_PROP_FPS))
    out = cv.VideoWriter(output_file, codec, fps, (output_width, output_height))

    # Track history
    track_history = defaultdict(lambda: [])
    track_class = defaultdict(lambda: [])
    class_count_list = [0, 0]
    count = 0

    while cap.isOpened():
        if count > max_count:
            print(f"Max count reached: {max_count}")
            break
        count += 1
        success, frame = cap.read()
        if not success:
            cap.release() # release object
            print("Video capture object not successful")
            break
        # skip frames based on FRAME_SKIP
        if count % frame_skip == 0:
            print(f'\rFrame: {count}', end='')
            sys.stdout.flush()
            count_str = '{:06d}'.format(count)
            image_name = video_filename + '_frame_' + count_str + '.jpg'
            save_path = os.path.join(output_dir, image_name)
            
            results = model.track(frame, persist=True, conf=conf_threshold, tracker=track_config, verbose=False)
            boxes = results[0].boxes.xywh.cpu()
            if len(boxes) != 0:
                track_class = save_track_img(results, track_history, track_class, frame, save_path, count)
            # Visualize the results on the frame
            annotated_frame = results[0].plot()
            out.write(annotated_frame)
            
            if cv.waitKey(1) & 0xFF == ord("q"):
               break # Break the loop if 'q' is pressed

    # Release the video capture object and close the output video writer
    cap.release()
    out.release()
    class_count_list = class_counts(track_class, class_count_list)
    # import code
    # code.interact(local=dict(globals(), **locals()))

    return len(track_history), class_count_list

def save_track_counts(manual_count_file, class_count_list, run_num, row_int, sheet_name):
    """Save the track counts against the manual counts
    Args:
        manual_count_file (str): Path to the manual count file.
        class_count_list (list): List of counts of each class
        run_num (int): Run number
        row_int (int): Row number in the excel file
    Returns:
        bool: True if successful"""
    wb = openpyxl.load_workbook(manual_count_file)
    sheet = wb[sheet_name]
    row = row_int+3
    col = int(run_num)*2+2
    sheet.cell(row=row, column=col).value = class_count_list[0]
    sheet.cell(row=row, column=col+1).value = class_count_list[1]
    wb.save(manual_count_file)
    wb.close()
    return True

def compute_stats(model_count, manual_count):
    """Compute TP, FP, FN and precision-recall metrics.
    Args:
        model_count (int): Count of objects detected by the model.
        manual_count (int): Count of objects manually annotated.
    Returns:
        tuple: TP, FP, FN, precision, recall"""
    if model_count > manual_count:
        FP = model_count - manual_count
        TP = manual_count
        FN = 0
    else:
        TP = model_count
        FP = 0
        FN = manual_count - model_count

    precision = 0 if (TP + FP) == 0 else TP / (TP + FP)
    recall = 0 if (TP + FN) == 0 else TP / (TP + FN)
    return TP, FP, FN, precision, recall

def eval(class_count_list, manual_count_file, video_path, run_num, sheet_name):
    """Print to terminal TP, FP, FN, precion and recall scores
    Args:
        class_count_list (list): List of counts of each class
        manual_count_file (str): Path to the manual count file.
        video_path (str): Path to the video file.
        run_num (int): Run number
    Returns:
        bool: True if successful"""
    man_tripy_count = None
    classstat = {class_label: {'TP': 0, 'FP': 0, 'FN': 0} for class_label in range(0, 2)}

    df = pd.read_excel(manual_count_file, sheet_name=sheet_name, skiprows=1)
    vid_count_list = df['Videos']
    vid_name = os.path.basename(video_path)[:-4]
    for i in range(0, len(vid_count_list)):
        if vid_count_list[i] == vid_name:
            man_tripy_count = df['Tripy'][i]
            man_rings_count = df['Rings'][i]
            break
    save_track_counts(manual_count_file, class_count_list, run_num, i, sheet_name)

    if man_tripy_count == None:
        print('Video counts not found')
        return False
    
    model_tripy_count, model_rings_count = class_count_list
    
    tripy_stats = compute_stats(model_tripy_count, man_tripy_count)
    print(f"Tripy class: TP={tripy_stats[0]}, FP={tripy_stats[1]}, FN={tripy_stats[2]}")
    print(f"Tripy class: Precision={tripy_stats[3]:.2f}, Recall={tripy_stats[4]:.2f}")

    # Compute TP, FP, FN and precision-recall metrics for Rings class
    rings_stats = compute_stats(model_rings_count, man_rings_count)
    print(f"Rings class: TP={rings_stats[0]}, FP={rings_stats[1]}, FN={rings_stats[2]}")
    print(f"Rings class: Precision={rings_stats[3]:.2f}, Recall={rings_stats[4]:.2f}")
    return True


def run():
    #for iteration over 1 video
    config = read_config("tracker/pipeline_config.yaml")
    video_path = config["video_path_in"]
    output_dir = config["save_dir"]
    weights_path = config["detection_model_path"]
    conf_threshold = config.get("detection_confidence_threshold")
    frame_skip = config.get("frame_skip")
    track_config = config.get("custom_tracker")
    model = YOLO(weights_path)
    sheet_name = config.get("sheet_name")
    run_num = config["num_times_run"]
    print("************")
    print(f"run_num:", run_num)
    print(f"Sheet name:", sheet_name) 
    print(f"Manual count file:", manual_count_file)
    print(f"Tracker_model:", track_config)
    print("************") 
    track_counts, class_count_list = track(video_path, output_dir, model, track_config, conf_threshold, frame_skip)
    print(f"Number of tracks: {track_counts}")
    print(f"Ring count: {class_count_list[0]} Tripys count: {class_count_list[1]}")
    manual_count_file = config.get("manual_count_file")
    eval(class_count_list, manual_count_file, video_path, run_num, sheet_name)
    print("Evaluation complete!")
 
## For bash script to iterate over multiple videos
video_path: str = sys.argv[1]
config = read_config("pipeline_config.yaml")
output_dir = config["save_dir"]
weights_path = config["detection_model_path"]
conf_threshold = config.get("detection_confidence_threshold")
frame_skip = config.get("frame_skip")
track_config = config.get("custom_tracker")
model = YOLO(weights_path)
manual_count_file = config.get("manual_count_file")
sheet_name = config.get("sheet_name")
run_num = config.get("num_times_run")
print("************")
print(f"run_num:", run_num)
print(f"Sheet name:", sheet_name) 
print(f"Manual count file:", manual_count_file)
print(f"Tracker_model:", track_config)
print("************")    
track_counts, class_count_list = track(video_path, output_dir, model, track_config, conf_threshold, frame_skip)
print(f"Number of tracks: {track_counts}")
print(f"Ring count: {class_count_list[0]} Tripys count: {class_count_list[1]}")
#eval(class_count_list, manual_count_file, video_path, run_num, sheet_name)
print(f"Evaluation complete for {video_path}")